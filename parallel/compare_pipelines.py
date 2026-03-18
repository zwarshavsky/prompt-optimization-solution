#!/usr/bin/env python3
"""
Compare results across all parallel optimization pipelines.

Reads Excel outputs from each pipeline, extracts per-cycle scores,
and produces a comparative summary showing which Gemini instruction
strategy performed best.

Usage:
  cd prompt-optimization-solution
  ./scripts/python/venv/bin/python3 parallel/compare_pipelines.py
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

PROJ_ROOT = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = PROJ_ROOT / "scripts" / "python" / "app_data" / "outputs"
STATUS_FILE = PROJ_ROOT / "parallel" / "status.json"

PIPELINE_LABELS = {
    "P1": "Original Simple (lean, no rules)",
    "P3": "Full Control (all rules, regression warning)",
    "P4": "Simplify-Every-Other (forced reduction on even cycles)",
    "P7": "Trend-Aware (score trajectory fed to Gemini)",
    "P10": "Gold Standard Start (pre-built 3K parser)",
}


def find_pipeline_excel(pipeline_id):
    """Find the Excel output file for a pipeline."""
    pattern = f"RiteHite_Opt_{pipeline_id}_*.xlsx"
    matches = sorted(OUTPUTS_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if matches:
        return matches[0]

    if STATUS_FILE.exists():
        try:
            status = json.loads(STATUS_FILE.read_text())
            p_info = status.get("pipelines", {}).get(pipeline_id, {})
            excel_name = p_info.get("excel_file")
            if excel_name:
                candidate = OUTPUTS_DIR / excel_name
                if candidate.exists():
                    return candidate
        except Exception:
            pass

    return None


def extract_scores_from_running_score(wb):
    """Extract per-cycle scores from the Running_Score sheet.

    The Running_Score sheet uses columns starting at col 4 (D) onward,
    with headers like 'Run 1', 'Run 2', etc. The Cycle number is in
    row 4 of each column. Metric labels are in column 1 (A).
    """
    if "Running_Score" not in wb.sheetnames:
        return []

    ws = wb["Running_Score"]
    cycles = []

    label_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label:
            label_rows[str(label).strip().lower()] = row_idx

    for col_idx in range(4, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if not header or ("Run" not in str(header) and "Cycle" not in str(header)):
            continue

        cycle_row = label_rows.get("cycle", 4)
        cycle_val = ws.cell(row=cycle_row, column=col_idx).value
        cycle_num = int(cycle_val) if cycle_val else len(cycles) + 1

        cycle_data = {"cycle": cycle_num, "pass": 0, "partial": 0, "fail": 0, "composite": 0}

        for row_idx in range(2, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            val = ws.cell(row=row_idx, column=col_idx).value

            if label and isinstance(val, (int, float)):
                label_lower = str(label).lower().strip()
                if label_lower == "pass":
                    cycle_data["pass"] = int(val)
                elif "partial" in label_lower:
                    cycle_data["partial"] = int(val)
                elif label_lower == "fail":
                    cycle_data["fail"] = int(val)

        cycle_data["composite"] = cycle_data["pass"] * 2 + cycle_data["partial"]
        cycles.append(cycle_data)

    return cycles


def extract_per_question_results(wb):
    """Extract per-question pass/partial/fail status across cycles."""
    results = {}
    for sheet_name in wb.sheetnames:
        cycle_match = re.search(r"cycle(\d+)", sheet_name, re.IGNORECASE)
        if not cycle_match:
            continue
        cycle_num = int(cycle_match.group(1))

        ws = wb[sheet_name]
        q_results = {}

        for row_idx in range(2, min(ws.max_row + 1, 50)):
            q_num = None
            status = None
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and re.match(r"^Q\d+$", str(val).strip()):
                    q_num = str(val).strip()
                if val and isinstance(val, str):
                    if "PASS" in val.upper() and "FAIL" not in val.upper():
                        status = "PASS"
                    elif "PARTIAL" in val.upper():
                        status = "PARTIAL"
                    elif "FAIL" in val.upper():
                        status = "FAIL"
            if q_num and status:
                q_results[q_num] = status

        if q_results:
            results[cycle_num] = q_results

    return results


def compare_all():
    """Run comparison across all pipeline outputs."""
    print("=" * 70)
    print("PARALLEL PIPELINE COMPARISON")
    print("=" * 70)
    print()

    pipeline_data = {}
    for pid in PIPELINE_LABELS:
        excel_path = find_pipeline_excel(pid)
        if excel_path and excel_path.exists():
            try:
                wb = load_workbook(str(excel_path), data_only=True)
                scores = extract_scores_from_running_score(wb)
                q_results = extract_per_question_results(wb)
                pipeline_data[pid] = {
                    "scores": scores,
                    "q_results": q_results,
                    "excel": excel_path.name,
                    "cycles": len(scores),
                }
                wb.close()
            except Exception as e:
                print(f"  {pid}: Error reading {excel_path.name}: {e}")
                pipeline_data[pid] = {"error": str(e)}
        else:
            print(f"  {pid}: No Excel output found")

    if not pipeline_data:
        print("No pipeline data found. Ensure pipelines have run and produced output.")
        return

    print("-" * 70)
    print("SCORE TRAJECTORY")
    print("-" * 70)
    print(f"{'Pipeline':<8} {'Strategy':<45} {'Cycles':<8} {'Best':<8} {'Final':<8}")
    print("-" * 70)

    rankings = []
    for pid in sorted(pipeline_data.keys()):
        data = pipeline_data[pid]
        if "error" in data:
            print(f"{pid:<8} {'ERROR':<45} {'-':<8} {'-':<8} {'-':<8}")
            continue

        scores = data["scores"]
        if not scores:
            print(f"{pid:<8} {PIPELINE_LABELS.get(pid, '')[:45]:<45} {'0':<8} {'-':<8} {'-':<8}")
            continue

        best = max(s["composite"] for s in scores)
        final = scores[-1]["composite"]
        cycles = len(scores)

        rankings.append((pid, best, final, cycles))
        label = PIPELINE_LABELS.get(pid, "")[:45]
        print(f"{pid:<8} {label:<45} {cycles:<8} {best:<8} {final:<8}")

    print()

    if rankings:
        print("-" * 70)
        print("COMPOSITE SCORE PER CYCLE")
        print("-" * 70)

        max_cycles = max(len(pipeline_data[pid]["scores"]) for pid in pipeline_data if "scores" in pipeline_data[pid])
        header = f"{'Pipeline':<8}" + "".join(f"{'C' + str(c+1):<8}" for c in range(max_cycles))
        print(header)
        print("-" * 70)

        for pid in sorted(pipeline_data.keys()):
            data = pipeline_data[pid]
            if "scores" not in data or not data["scores"]:
                continue
            row = f"{pid:<8}"
            for c in range(max_cycles):
                if c < len(data["scores"]):
                    row += f"{data['scores'][c]['composite']:<8}"
                else:
                    row += f"{'-':<8}"
            print(row)

        print()
        print("-" * 70)
        print("PASS/PARTIAL/FAIL BREAKDOWN PER CYCLE")
        print("-" * 70)

        for pid in sorted(pipeline_data.keys()):
            data = pipeline_data[pid]
            if "scores" not in data or not data["scores"]:
                continue
            print(f"\n  {pid} ({PIPELINE_LABELS.get(pid, '')}):")
            for s in data["scores"]:
                bar_pass = "█" * s["pass"]
                bar_partial = "▒" * s["partial"]
                bar_fail = "░" * s["fail"]
                print(f"    C{s['cycle']}: {bar_pass}{bar_partial}{bar_fail}  P={s['pass']} Pt={s['partial']} F={s['fail']} (composite={s['composite']})")

    if rankings:
        print()
        print("-" * 70)
        print("RANKINGS (by best composite score)")
        print("-" * 70)
        rankings.sort(key=lambda x: (-x[1], -x[2]))
        for rank, (pid, best, final, cycles) in enumerate(rankings, 1):
            trend = "↑" if final >= best else "↓"
            print(f"  #{rank}: {pid} — best={best}, final={final} {trend}, cycles={cycles} — {PIPELINE_LABELS.get(pid, '')}")

        winner = rankings[0]
        print(f"\n  WINNER: {winner[0]} with composite score {winner[1]}")
        print(f"  Strategy: {PIPELINE_LABELS.get(winner[0], '')}")

    summary = {
        "generated_at": str(Path(__file__).name),
        "timestamp": __import__("datetime").datetime.now().isoformat(),
        "rankings": [
            {"rank": i + 1, "pipeline": r[0], "best_composite": r[1], "final_composite": r[2], "cycles": r[3], "strategy": PIPELINE_LABELS.get(r[0], "")}
            for i, r in enumerate(rankings)
        ] if rankings else [],
    }
    summary_file = PROJ_ROOT / "parallel" / "comparison_results.json"
    with open(summary_file, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"\nResults saved to: {summary_file}")


if __name__ == "__main__":
    compare_all()
