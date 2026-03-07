"""
Excel I/O helpers.

create_analysis_sheet_with_prompts is moved here from test_gemini.py to centralize
Excel-related logic.
"""

import sys
import json
from pathlib import Path
import pandas as pd
from salesforce_api import get_salesforce_credentials, invoke_prompt, retrieve_metadata_via_api, SearchIndexAPI
import xml.etree.ElementTree as ET
import xml.etree.ElementTree as ET

def log_print(*args, **kwargs):
    """Print with immediate flush for real-time output"""
    print(*args, **kwargs, flush=True)


def get_input_column_headers_and_rows(config_dict):
    """
    Single source of truth for prompt input columns used in both the analysis sheet and Running_Score.
    Returns headers and per-question row values so first tab and run tabs stay in sync.

    Args:
        config_dict: Full config (e.g. YAML). May contain optional 'promptInputs' and 'questions'.

    Returns:
        tuple: (input_headers, list_of_rows)
            - input_headers: e.g. ['Q#', 'Question'] or ['Q#', 'Product', 'Question']
            - list_of_rows: list of lists, one per question, same length as input_headers
    """
    config = config_dict or {}
    # Support full YAML: questions at top level, promptInputs under configuration
    questions = config.get("questions") or []
    prompt_inputs = config.get("promptInputs") or (config.get("configuration") or {}).get("promptInputs") or []

    if not prompt_inputs:
        # Single-input: Q# + Question
        headers = ["Q#", "Question"]
        rows = [[q.get("number", ""), q.get("text", "")] for q in questions]
        return headers, rows

    # Multi-input: Q# + one column per promptInputs (by displayName)
    headers = ["Q#"] + [p.get("displayName") or p.get("apiName", "") for p in prompt_inputs]
    api_names = [p.get("apiName", "") for p in prompt_inputs]

    def row_for(q):
        vals = [q.get("number", "")]
        inputs_map = q.get("inputs") or {}
        for api_name in api_names:
            val = inputs_map.get(api_name)
            if val is None and api_name == "Input:Question":
                val = q.get("text", "")
            vals.append(val if val is not None else "")
        return vals

    rows = [row_for(q) for q in questions]
    return headers, rows


def create_analysis_sheet_with_prompts(excel_file, questions_list=None,
                                      prompt_template_name=None, search_index_id=None, models_list=None,
                                      refinement_stage=None, cycle_number=None, config_dict=None):
    """
    Create a new timestamped sheet with columns A-D, get prompts/metadata, and fetch answers from Salesforce
    """
    # #region agent log (optional - only if debug.log exists locally)
    try:
        debug_log_path = Path('/Users/zwarshavsky/Documents/Custom_LWC_Org_SDO/Custom LWC Development SDO/.cursor/debug.log')
        if debug_log_path.parent.exists():
            with open(debug_log_path, 'a') as f:
                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"ALL","location":"excel_io.py:create_analysis_sheet_with_prompts","message":"ENTRY","data":{"excel_file":str(excel_file),"has_questions_list":questions_list is not None,"questions_count":len(questions_list) if questions_list else 0,"prompt_template":prompt_template_name},"timestamp":int(__import__('time').time()*1000)}) + '\n')
    except (FileNotFoundError, OSError):
        # Silently skip debug logging if path doesn't exist (e.g., on Heroku)
        pass
    # #endregion

    if not prompt_template_name:
        log_print("❌ Error: --prompt-template is required")
        sys.exit(1)
    if not search_index_id:
        log_print("❌ Error: --search-index-id is required")
        sys.exit(1)
    if pd is None:
        log_print("❌ pandas is required for Excel file processing")
        sys.exit(1)

    # CRITICAL: If run_id is provided and file doesn't exist on disk, try loading from database first
    # This prevents overwriting existing Excel files that are stored in DB but not on ephemeral filesystem
    run_id_from_config = None
    if config_dict:
        # Try to extract run_id from config_dict (passed from main.py workflow)
        run_id_from_config = config_dict.get('_run_id') or config_dict.get('run_id')
    
    # Resolve Excel file path - handle both absolute and relative paths
    excel_path = Path(excel_file)
    if not excel_path.is_absolute():
        if not excel_path.exists():
            script_dir = Path(__file__).parent.parent
            excel_path_attempt = script_dir / excel_file
            if excel_path_attempt.exists():
                excel_path = excel_path_attempt
            else:
                if "prompt-optimization-solution" in excel_file:
                    relative_part = excel_file.replace("prompt-optimization-solution/", "")
                    excel_path_attempt = script_dir / relative_part
                    if excel_path_attempt.exists():
                        excel_path = excel_path_attempt
                    else:
                        inputs_dir = script_dir / "inputs"
                        excel_path_attempt = inputs_dir / Path(excel_file).name
                        if excel_path_attempt.exists():
                            excel_path = excel_path_attempt

    # CRITICAL FIX: If file doesn't exist on disk but run_id is provided, try loading from database
    # This prevents overwriting existing Excel files on Heroku (ephemeral filesystem)
    if not excel_path.exists() and run_id_from_config and "run_" in str(excel_path):
        try:
            from app import load_excel_from_db
            loaded_path = load_excel_from_db(run_id_from_config)
            if loaded_path and Path(loaded_path).exists():
                excel_path = Path(loaded_path)
                log_print(f"   ✅ Loaded Excel file from database: {excel_path.name}")
            else:
                log_print(f"   ℹ️  Excel file not in database, will create new: {excel_path.name}")
        except Exception as e:
            log_print(f"   ⚠️  Could not load from database: {e}, will create new file")

    # For new runs, the Excel file may not exist yet - that's OK, we'll create it
    if not excel_path.exists():
        if "run_" in str(excel_path) and excel_path.suffix == '.xlsx':
            log_print(f"   ℹ️  New Excel file will be created: {excel_path.name}")
        else:
            log_print(f"⚠️  Excel file not found: {excel_file}")
            log_print(f"   Tried: {excel_path}")
            log_print(f"   Will attempt to create if this is a new run")

    excel_file = str(excel_path)  # Update to absolute path

    log_print("="*80)
    log_print("Creating New Analysis Sheet with Prompts")
    log_print("="*80)
    log_print(f"📄 Excel File: {excel_file}")
    log_print(f"🔮 Prompt Template: {prompt_template_name}")
    log_print(f"🔍 Search Index ID: {search_index_id}")
    log_print()

    # Get questions list - must be provided
    if questions_list is None or len(questions_list) == 0:
        log_print("❌ Error: questions_list must be provided")
        sys.exit(1)

    log_print(f"📋 Using provided questions list: {len(questions_list)} questions")
    # Normalize: support tuple (q_num, q_text, expected_answer) or dict (multi-input)
    normalized_questions = []
    for q in questions_list:
        if isinstance(q, dict):
            normalized_questions.append(q)
        elif len(q) == 2:
            normalized_questions.append((q[0], q[1], ''))
        elif len(q) == 3:
            normalized_questions.append(q)
        else:
            log_print(f"   ⚠️  Skipping invalid question format: {q}")
    questions_list = normalized_questions

    # Input columns from config (single source of truth for first tab and Running_Score)
    input_headers, input_rows = get_input_column_headers_and_rows(config_dict)
    outcome_columns = ['Received Answer', 'Expected Answer', 'Model Used', 'Pass/Fail', 'Safety Score', 'Root Cause/Explanation', 'Prompt Modification Next Version']
    all_columns = input_headers + outcome_columns
    num_input_cols = len(input_headers)

    # Build DataFrame: use config-derived input rows and expected_answer from questions_list by index
    def expected_for(q):
        return q.get('expectedAnswer', '') if isinstance(q, dict) else (q[2] if len(q) >= 3 else '')

    header_data = [all_columns]
    question_data = []
    for i in range(len(questions_list)):
        input_row = input_rows[i] if i < len(input_rows) else [''] * num_input_cols
        exp = expected_for(questions_list[i])
        question_data.append(input_row + ['', exp, '', '', '', '', ''])
    df_new = pd.DataFrame(header_data + question_data, columns=all_columns)

    # Metadata rows at top (matching V7 format)
    log_print("   🔑 Getting Salesforce credentials for metadata...")
    instance_url, access_token = get_salesforce_credentials(config_dict=config_dict)
    log_print(f"   ✅ Connected: {instance_url}")
    org_username = config_dict.get('configuration', {}).get('salesforce', {}).get('username', 'Unknown') if config_dict else "Unknown"
    search_index_label = "Unknown"
    try:
        log_print(f"   🔍 Retrieving search index label for ID: {search_index_id}...")
        search_api = SearchIndexAPI(instance_url, access_token)
        full_index = search_api.get_index(search_index_id)
        if full_index:
            search_index_label = full_index.get('label', 'Unknown')
            log_print(f"   ✅ Found index label: {search_index_label}")
        else:
            log_print(f"   ⚠️  Index not found")
    except Exception as e:
        log_print(f"   ⚠️  Could not retrieve index label: {e}")
        pass

    num_cols = len(df_new.columns)
    metadata_rows = [
        ['Environment:', instance_url, org_username, ''] + [''] * (num_cols - 4),
        ['Search Index:', search_index_label] + [''] * (num_cols - 2),
        ['Prompt Builder Prompt:', prompt_template_name] + [''] * (num_cols - 2),
        ['Stage Status:', ''] + [''] * (num_cols - 2),
        ['Stage Status Reason:', ''] + [''] * (num_cols - 2),
        [''] * num_cols,
        [''] * num_cols,
    ]
    metadata_header = pd.DataFrame(metadata_rows, columns=df_new.columns)

    log_print(f"\n🔮 Invoking Salesforce prompt '{prompt_template_name}' (API name) for each question...")
    log_print("   🔑 Getting Salesforce credentials...")
    instance_url, access_token = get_salesforce_credentials(config_dict=config_dict)
    log_print(f"   ✅ Connected to: {instance_url}")

    # Model from prompt template (display only)
    response_model = ""
    try:
        log_print(f"   📋 Retrieving prompt template metadata for: {prompt_template_name}...")
        dev_name = prompt_template_name
        xml_content = retrieve_metadata_via_api(instance_url, access_token, "GenAiPromptTemplate", dev_name)
        if xml_content:
            root = ET.fromstring(xml_content)
            ns = {'met': 'http://soap.sforce.com/2006/04/metadata'}
            active_version = root.find('.//met:activeVersionIdentifier', ns)
            if active_version is not None:
                active_id = active_version.text
                for version in root.findall('.//met:templateVersions', ns):
                    version_id_elem = version.find('met:versionIdentifier', ns)
                    if version_id_elem is not None and version_id_elem.text == active_id:
                        model_elem = version.find('met:primaryModel', ns)
                        if model_elem is not None:
                            response_model = model_elem.text or ''
                            log_print(f"   ✅ Found model in prompt template: {response_model} (for display only)")
                            break
    except Exception as e:
        log_print(f"   ⚠️  Could not retrieve model: {e}")
        response_model = "Unknown"

    if not models_list or len(models_list) == 0:
        log_print(f"   ⚠️  WARNING: No models_list provided from YAML config! API calls may fail.")
    else:
        log_print(f"   ✅ Using models from YAML config: {models_list[0]} (primary) + {len(models_list)-1} fallback(s)")

    log_print(f"   📝 Processing {len(questions_list)} questions...")
    for q in questions_list:
        q_num = q.get('number', '') if isinstance(q, dict) else (q[0] if q else '')
        q_preview = (q.get('text', '') or list((q.get('inputs') or {}).values())[:1] or [''])[0] if isinstance(q, dict) else (q[1][:60] if len(q) > 1 else '')
        log_print(f"   ✓ {q_num}: {str(q_preview)[:60]}...")

    prompt_dev_name = prompt_template_name
    log_print(f"   🔧 Using prompt API name (DeveloperName): {prompt_dev_name}")

    # Log GenAI prompt template active version and retriever (for traceability)
    genai_prompt_active_version = None
    genai_prompt_retriever = None
    try:
        xml_content = retrieve_metadata_via_api(instance_url, access_token, "GenAiPromptTemplate", prompt_dev_name)
        if xml_content:
            root = ET.fromstring(xml_content)
            ns = {'met': 'http://soap.sforce.com/2006/04/metadata'}
            active_el = root.find('.//met:activeVersionIdentifier', ns)
            genai_prompt_active_version = active_el.text if active_el is not None and active_el.text else None
            if genai_prompt_active_version:
                for v in root.findall('.//met:templateVersions', ns):
                    vid_el = v.find('met:versionIdentifier', ns)
                    if vid_el is not None and vid_el.text == genai_prompt_active_version:
                        for tdp in v.findall('met:templateDataProviders', ns):
                            def_el = tdp.find('met:definition', ns)
                            if def_el is not None and def_el.text and 'getEinsteinRetrieverResults' in def_el.text:
                                genai_prompt_retriever = def_el.text.split('/')[-1] if '/' in def_el.text else def_el.text
                                break
                        break
            log_print(f"   📌 GenAI prompt active version: {genai_prompt_active_version or 'unknown'}")
            log_print(f"   📌 GenAI prompt retriever: {genai_prompt_retriever or 'unknown'}")
        else:
            log_print(f"   ⚠️  Could not retrieve prompt template metadata for version logging")
    except Exception as e:
        log_print(f"   ⚠️  Could not log GenAI prompt version: {e}")

    total = len(questions_list)
    all_responses = []
    log_print(f"\n   🚀 Starting prompt invocations (sequential, {total} questions)...")
    for idx, q in enumerate(questions_list):
        q_num = q.get('number', '') if isinstance(q, dict) else (q[0] if q else '')
        log_print(f"\n   🔄 [{idx+1}/{total}] Invoking prompt for {q_num}...")
        # Step 6: pass question and/or input_value_map (handled in next edit for invoke_prompt call)
        q_text = q.get('text', '') if isinstance(q, dict) else (q[1] if len(q) > 1 else '')
        input_value_map = (q.get('inputs') or None) if isinstance(q, dict) else None
        if input_value_map and not any(input_value_map.values()):
            input_value_map = None
        if input_value_map:
            log_print(f"      Inputs: {list(input_value_map.keys())}")
        else:
            log_print(f"      Question: {str(q_text)[:80]}...")
        try:
            result, model_used = invoke_prompt(
                instance_url, access_token, q_text or None, prompt_dev_name,
                max_retries=3, model_used=None, models_list=models_list,
                run_id=config_dict.get('_run_id') or config_dict.get('run_id') if config_dict else None,
                input_value_map=input_value_map
            )
            
            # CRITICAL: Check if job was killed/aborted
            if model_used == "ABORTED":
                run_id = config_dict.get('_run_id') or config_dict.get('run_id') if config_dict else None
                log_print(f"   🛑 ABORTING: Job {run_id} was killed during prompt invocation. Stopping workflow.", flush=True)
                raise RuntimeError(f"Job {run_id} aborted during prompt invocation.")
            
            # Handle None result (shouldn't happen normally, but be safe)
            if result is None:
                result = "No response received"
            
            log_print(f"      ✅ Response received (model: {model_used}, length: {len(result)} chars)")
            if 'Error' in result or 'API Error' in result:
                log_print(f"   ⚠️  {q_num}: Error - {result[:100]}")
            else:
                log_print(f"   ✅ {q_num}: Success")
        except RuntimeError:
            # Re-raise RuntimeError (abort) to stop the workflow
            raise
        except Exception as e:
            result = f"Error invoking prompt: {e}"
            model_used = "Unknown"
            log_print(f"   ⚠️  {q_num}: Exception - {e}")
        all_responses.append((q_num, result, model_used))

    log_print("   📝 Updating DataFrame with answers...")
    df_answers = df_new.copy()
    col_received = num_input_cols  # Received Answer
    col_model = num_input_cols + 2  # Model Used
    for i, (q_num, answer_text, model_used) in enumerate(all_responses):
        if i < len(df_answers):
            df_answers.iloc[i + 1, col_received] = answer_text
            df_answers.iloc[i + 1, col_model] = model_used

    # Build final DataFrame: metadata + answers
    final_df = pd.concat([metadata_header, df_answers], ignore_index=True)

    # Get LLM parser prompt from search index
    llm_parser_prompt_current = ""
    try:
        log_print(f"   🔍 Retrieving LLM parser prompt from search index...")
        search_api = SearchIndexAPI(instance_url, access_token)
        full_index = search_api.get_index(search_index_id)
        if full_index:
            parsing_configs = full_index.get('parsingConfigurations', [])
            for config in parsing_configs:
                config_obj = config.get('config', {})
                config_id = config_obj.get('id', '').lower()
                if 'llm' in config_id or 'parse_documents_using_llm' in config_id:
                    user_values = config_obj.get('userValues', [])
                    for uv in user_values:
                        if uv.get('id') == 'prompt':
                            llm_parser_prompt_current = uv.get('value', '')
                            if llm_parser_prompt_current:
                                log_print(f"   ✅ Found LLM parser prompt ({len(llm_parser_prompt_current)} chars)")
                                break
                    if llm_parser_prompt_current:
                        break
            if not llm_parser_prompt_current:
                log_print(f"   ⚠️  LLM parser prompt not found in index")
        else:
            log_print(f"   ⚠️  Search index not found")
    except Exception as e:
        log_print(f"   ⚠️  Could not retrieve LLM parser prompt: {e}")

    # Get Prompt Builder Prompt from prompt template metadata
    prompt_builder_prompt_current = ""
    try:
        log_print(f"   🔍 Retrieving Prompt Builder Prompt from template metadata...")
        xml_content = retrieve_metadata_via_api(instance_url, access_token, "GenAiPromptTemplate", prompt_template_name)
        if xml_content:
            root = ET.fromstring(xml_content)
            ns = {'met': 'http://soap.sforce.com/2006/04/metadata'}
            active_version = root.find('.//met:activeVersionIdentifier', ns)
            if active_version is not None:
                active_id = active_version.text
                for version in root.findall('.//met:templateVersions', ns):
                    version_id_elem = version.find('met:versionIdentifier', ns)
                    if version_id_elem is not None and version_id_elem.text == active_id:
                        content_elem = version.find('met:content', ns)
                        if content_elem is not None and content_elem.text:
                            prompt_builder_prompt_current = content_elem.text.strip()
                            # Unescape HTML entities
                            prompt_builder_prompt_current = prompt_builder_prompt_current.replace('&quot;', '"').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
                            log_print(f"   ✅ Found Prompt Builder Prompt ({len(prompt_builder_prompt_current)} chars)")
                            break
            if not prompt_builder_prompt_current:
                log_print(f"   ⚠️  Prompt Builder Prompt not found in template")
        else:
            log_print(f"   ⚠️  Could not retrieve template metadata")
    except Exception as e:
        log_print(f"   ⚠️  Could not retrieve Prompt Builder Prompt: {e}")

    # Add bottom metadata (parser/prompt, etc.) - match old working structure
    # Each label has its value in the NEXT row, column 1 (B)
    num_cols = len(final_df.columns)
    
    # Determine stage-specific values
    if refinement_stage == "agentforce_agent":
        agentforce_config_current_value = 'Not found'  # TODO: retrieve if needed
        agentforce_config_proposed_value = ''  # Empty, will be filled by Gemini
    else:
        agentforce_config_current_value = 'not the current stage'
        agentforce_config_proposed_value = 'not the current stage'
    
    if refinement_stage == "response_prompt":
        prompt_builder_proposed_value = ''  # Empty, will be filled by Gemini
    else:
        prompt_builder_proposed_value = 'not the current stage'
    
    bottom_rows = [
        [''] * num_cols,
        ['GenAI Prompt Active Version:', ''] + [''] * (num_cols - 2),
        ['', (genai_prompt_active_version or 'unknown')] + [''] * (num_cols - 2),
        [''] * num_cols,
        ['GenAI Prompt Retriever:', ''] + [''] * (num_cols - 2),
        ['', (genai_prompt_retriever or 'unknown')] + [''] * (num_cols - 2),
        [''] * num_cols,
        ['LLM Parser Prompt Current:', ''] + [''] * (num_cols - 2),
        ['', llm_parser_prompt_current[:50000] if llm_parser_prompt_current else 'Not found'] + [''] * (num_cols - 2),  # Value in next row, column 1
        [''] * num_cols,
        ['LLM Parser Prompt Proposed from Gemini:', ''] + [''] * (num_cols - 2),
        ['', ''] + [''] * (num_cols - 2),  # Empty, will be filled by Gemini
        [''] * num_cols,
        ['Prompt Builder Prompt:', ''] + [''] * (num_cols - 2),
        ['', prompt_builder_prompt_current[:50000] if prompt_builder_prompt_current else 'Not found'] + [''] * (num_cols - 2),  # Always populated with current
        [''] * num_cols,
        ['Prompt Builder Prompt Proposed from Gemini:', ''] + [''] * (num_cols - 2),
        ['', prompt_builder_proposed_value] + [''] * (num_cols - 2),  # "not the current stage" or empty
        [''] * num_cols,
        ['Agentforce Agent Configuration Current:', ''] + [''] * (num_cols - 2),
        ['', agentforce_config_current_value] + [''] * (num_cols - 2),
        [''] * num_cols,
        ['Agentforce Agent Configuration Proposed from Gemini:', ''] + [''] * (num_cols - 2),
        ['', agentforce_config_proposed_value] + [''] * (num_cols - 2),
        [''] * num_cols,
        ['Instructions to Gemini:', ''] + [''] * (num_cols - 2),
        ['', ''] + [''] * (num_cols - 2)  # Empty initially, will be filled with full instructions after Gemini analysis
    ]
    bottom_df = pd.DataFrame(bottom_rows, columns=final_df.columns)

    final_df = pd.concat([final_df, bottom_df], ignore_index=True)

    # Write to Excel
    # CRITICAL FIX: Always use append mode ('a') if file exists, even if just loaded from DB
    # Never use 'w' mode for run-specific files as it would overwrite existing cycles
    from openpyxl import load_workbook
    
    # CRITICAL: Ensure directory exists before writing
    excel_path = Path(excel_file)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    excel_file_exists = excel_path.exists()
    
    try:
        # CRITICAL FIX: pd.ExcelWriter mode='a' REQUIRES the file to exist
        # Only use append mode if file actually exists on disk
        # Use write mode to create new files (even for run files)
        if excel_file_exists:
            # File exists - use append mode to preserve existing sheets
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                final_df.to_excel(writer, sheet_name=f"analysis_{refinement_stage}_cycle{cycle_number}_{Path(excel_file).stem}", index=False, header=False)
        else:
            # File doesn't exist - use write mode to create it
            # This handles both new run files and files that were in DB but not on disk
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
                final_df.to_excel(writer, sheet_name=f"analysis_{refinement_stage}_cycle{cycle_number}_{Path(excel_file).stem}", index=False, header=False)
    except Exception as e:
        error_msg = f"CRITICAL: Cannot write Excel file to {excel_file}: {e}"
        log_print(f"❌ {error_msg}")
        # Raise exception instead of sys.exit so worker can catch and mark job as failed
        raise RuntimeError(error_msg) from e

    # Ensure Running_Score sheet exists (create if new file or if missing)
    # Match old working structure exactly
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file)
        if 'Running_Score' not in wb.sheetnames:
            # Create new sheet at position 0 (first tab); use same input columns as analysis sheet
            ws = wb.create_sheet('Running_Score', 0)
            run_input_headers, run_input_rows = get_input_column_headers_and_rows(config_dict)
            num_input_cols_run = len(run_input_headers)
            first_run_col = num_input_cols_run + 1

            rows = []
            # Row 1: Headers (Run 1 in column first_run_col)
            header_row = ['Metric', ''] + [''] * (num_input_cols_run - 1) + ['Run 1']
            rows.append(header_row)
            # Rows 2-11: Summary metrics (label in col 1, value will go in first_run_col)
            for label in ['Run ID', 'Timestamp', 'Cycle', 'Pass', 'Partial', 'Fail', 'Total', 'Pass Rate', 'Avg Safety', 'Stage Status']:
                rows.append([label, ''] + [''] * (num_input_cols_run - 1))
            rows.append([''])  # Empty separator
            # Question block: same input column headers + empty for Run 1
            question_header = run_input_headers + ['']
            rows.append(question_header)
            for input_row in run_input_rows:
                rows.append(input_row + [''])
            
            # Write initial structure
            for row_idx, row_data in enumerate(rows, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(excel_file)
            wb.close()
            log_print(f"   ✅ Created Running_Score sheet")
        else:
            wb.close()
    except Exception as e:
        log_print(f"   ⚠️  Could not create/verify Running_Score sheet: {e}")

    sheet_name = f"analysis_{refinement_stage}_cycle{cycle_number}_{Path(excel_file).stem}"
    log_print(f"   ✅ Created new sheet: {sheet_name}")
    return sheet_name


def update_run_summary_sheet(excel_file, run_id, cycle_number, results_data, config_dict):
    """
    Update or create Running_Score sheet with cycle results.
    This sheet tracks all cycles across all runs, with each run/cycle as a column.
    
    Args:
        excel_file: Path to run-specific Excel file
        run_id: Unique run identifier
        cycle_number: Current cycle number
        results_data: Dict with cycle results containing:
            - timestamp: str
            - pass_count: int
            - fail_count: int
            - total: int
            - pass_rate: float
            - avg_safety: float
            - stage_status: str
            - question_results: list of dicts with 'q_number' and 'status'
        config_dict: Frozen YAML config (for questions list)
    """
    import openpyxl
    from pathlib import Path
    
    sheet_name = 'Running_Score'
    excel_path = Path(excel_file)
    
    if not excel_path.exists():
        log_print(f"  ⚠️  Excel file not found: {excel_file}")
        return
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Input columns from config (same as first tab and Running_Score create)
        run_input_headers, run_input_rows = get_input_column_headers_and_rows(config_dict)
        num_input_cols_run = len(run_input_headers)
        first_run_col = num_input_cols_run + 1

        # Check if Running_Score sheet exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing_runs = True
        else:
            # Create new sheet at position 0 (first tab); same structure as in create_analysis_sheet
            ws = wb.create_sheet(sheet_name, 0)
            existing_runs = False

        # Determine column index for new run/cycle
        if existing_runs:
            last_col = first_run_col
            while ws.cell(row=2, column=last_col).value is not None and str(ws.cell(row=2, column=last_col).value).strip():
                last_col += 1
            new_col_index = last_col
        else:
            new_col_index = first_run_col
            rows = []
            header_row = ['Metric', ''] + [''] * (num_input_cols_run - 1) + ['Run 1']
            rows.append(header_row)
            for label in ['Run ID', 'Timestamp', 'Cycle', 'Pass', 'Partial', 'Fail', 'Total', 'Pass Rate', 'Avg Safety', 'Stage Status']:
                rows.append([label, ''] + [''] * (num_input_cols_run - 1))
            rows.append([''])
            question_header = run_input_headers + ['']
            rows.append(question_header)
            for input_row in run_input_rows:
                rows.append(input_row + [''])
            for row_idx, row_data in enumerate(rows, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Update header for this column (Run 1 when new_col_index == first_run_col, etc.)
        run_label = f"Run {new_col_index - num_input_cols_run}"
        ws.cell(row=1, column=new_col_index, value=run_label)
        
        # Populate summary metrics (rows 2-11)
        ws.cell(row=2, column=new_col_index, value=run_id)
        ws.cell(row=3, column=new_col_index, value=results_data['timestamp'])
        ws.cell(row=4, column=new_col_index, value=cycle_number)
        ws.cell(row=5, column=new_col_index, value=results_data['pass_count'])
        ws.cell(row=6, column=new_col_index, value=results_data.get('partial_count', 0))
        ws.cell(row=7, column=new_col_index, value=results_data['fail_count'])
        ws.cell(row=8, column=new_col_index, value=results_data['total'])
        ws.cell(row=9, column=new_col_index, value=f"{results_data['pass_rate']:.1f}%")
        ws.cell(row=10, column=new_col_index, value=results_data['avg_safety'])
        ws.cell(row=11, column=new_col_index, value=results_data['stage_status'])
        
        # Populate question results (starting at row 13, after header row 12)
        question_start_row = 13
        for q_result in results_data['question_results']:
            q_number = q_result['q_number']
            status = q_result['status']
            
            # Find row for this question
            for row_idx in range(question_start_row, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value == q_number:
                    if status == 'PASS':
                        cell_value = '✅ PASS'
                    elif status == 'PARTIAL':
                        cell_value = '🔶 PARTIAL'
                    else:
                        cell_value = '❌ FAIL'
                    ws.cell(row=row_idx, column=new_col_index, value=cell_value)
                    break
        
        # Ensure Running_Score is first sheet
        if wb.sheetnames[0] != sheet_name:
            wb.move_sheet(wb[sheet_name], offset=-len(wb.sheetnames))
        
        # Save workbook
        wb.save(excel_file)
        wb.close()
        
        log_print(f"  ✅ Updated Running_Score sheet with Run {new_col_index - 2} (Cycle {cycle_number})")
        
    except Exception as e:
        log_print(f"  ⚠️  Error updating Running_Score sheet: {e}")
        import traceback
        traceback.print_exc()


__all__ = ["create_analysis_sheet_with_prompts", "update_run_summary_sheet"]



